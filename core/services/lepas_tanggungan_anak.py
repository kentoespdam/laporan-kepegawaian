import pandas as pd
from openpyxl import load_workbook

from core.excel_helper import font_style, text_align, write_data_to_excel, save_workbook
from core.helper import format_date_vectorized
from core.model.lepas_tanggungan_anak import (
    FilterLepasTanggunganAnak,
    fetch_lepas_tanggungan_anak,
)


def data_lepas_tanggungan_anak(filter: FilterLepasTanggunganAnak = FilterLepasTanggunganAnak.BULAN_INI):
    data = fetch_lepas_tanggungan_anak(filter)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(df: pd.DataFrame):
    df["tanggal_lahir"] = format_date_vectorized(df["tanggal_lahir"])
    df["tanggungan"] = df["tanggungan"].eq(1)
    return df


def to_excel(title_text: str, filter: FilterLepasTanggunganAnak = FilterLepasTanggunganAnak.BULAN_INI):
    data = data_lepas_tanggungan_anak(filter)
    if data.empty:
        return None

    wb = load_workbook("template/template_lta.xlsx")
    ws = wb.active

    title_cell = ws.cell(row=2, column=1)
    title_cell.value = title_text
    title_cell.alignment = text_align("center")
    title_cell.font = font_style("bold")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    column_mappings = [
        ("index", lambda idx, _: int(idx + 1), ["bold", "allborder"]),
        ("nama_anak", lambda _, row: row["nama_anak"], ["allborder"]),
        ("jenis_kelamin", lambda _, row: row["jenis_kelamin"], ["allborder"]),
        ("tanggal_lahir", lambda _, row: row["tanggal_lahir"], ["allborder"]),
        ("umur", lambda _, row: row["umur"], ["allborder"]),
        ("status_pendidikan", lambda _, row: row["status_pendidikan"], ["allborder"]),
        ("nama_karyawan", lambda _, row: row["nama_karyawan"], ["allborder"]),
        ("nipam", lambda _, row: row["nipam"], ["allborder"]),
        ("nama_jabatan", lambda _, row: row["nama_jabatan"], ["allborder"]),
    ]

    # Write data to Excel
    write_data_to_excel(ws, data, column_mappings, 5)

    return save_workbook(wb)
