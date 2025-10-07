import io
import itertools

import pandas as pd
from openpyxl import load_workbook

from core.excel_helper import cell_builder, save_workbook
from core.helper import hitung_sisa_bulan
from core.model.dnp import fetch_dnp
from core.model.organisasi import fetch_kode_nama_organisasi


def fetch_dnp_data() -> dict[str, pd.DataFrame]:
    """Fetch DNP data from database and return it in a standardized format"""
    dnp_data = fetch_dnp()
    dnp_data = _cleanup_dnp_data(dnp_data)
    organisasi = _fetch_organisasi()
    return {
        "dnp": dnp_data,
        "organisasi": organisasi
    }


def _cleanup_dnp_data(dnp_data: pd.DataFrame) -> pd.DataFrame:
    """Clean up DNP data"""
    dnp_data["mkg_tahun"] = dnp_data["mkg_tahun"].fillna(0).clip(lower=0).astype(int)
    dnp_data["mkg_bulan"] = hitung_sisa_bulan(dnp_data["mkg_tahun"], dnp_data["mkg_bulan"])
    dnp_data["mk_bulan"] = hitung_sisa_bulan(dnp_data["mk_tahun"], dnp_data["mk_bulan"])

    # cleanup organisasi
    mask = dnp_data["level_jabatan"].isin([2, 3, 4])
    dnp_data["kode_organisasi"] = dnp_data["kode_organisasi"].where(~mask, "1")
    return dnp_data


def _fetch_organisasi() -> pd.DataFrame:
    """Fetch and return organisasi data"""
    organisasi = fetch_kode_nama_organisasi(4)
    direksi_org = pd.DataFrame({"kode": ["1"], "nama": ["DIREKSI"]})
    organisasi = pd.concat([direksi_org, organisasi])
    return organisasi


def to_excel(tahun: int, bulan: int) -> io.BytesIO:
    """Generate Excel file from DNP data"""
    df = fetch_dnp_data()
    wb = load_workbook('template/template_dnp.xlsx')
    ws = wb.active

    # Set title
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = f"BULAN : {bulan} {tahun}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)

    # Pre-cache data
    dnp_df = df["dnp"]
    organisasi_df = df["organisasi"]

    # Optimasi: buat dictionary untuk lookup pegawai
    pegawai_by_org = _build_pegawai_lookup(dnp_df)

    row_num = itertools.count(start=8)
    root_urut = itertools.count(start=1)

    for org_row in organisasi_df.itertuples():
        # Header organisasi
        curr_row_num = next(row_num)
        cell_builder(ws, curr_row_num, 1, org_row.nama, ["bold", "vborder"])
        ws.merge_cells(start_row=curr_row_num, start_column=1,
                       end_row=curr_row_num, end_column=16)

        # Data pegawai
        child_urut = itertools.count(start=1)
        pegawai_data = pegawai_by_org.get(org_row.kode, pd.DataFrame())

        for peg in pegawai_data.itertuples():
            _add_pegawai_row(ws, next(row_num), next(root_urut), next(child_urut), peg)

        # Empty row separator
        curr_row_num = next(row_num)
        cell_builder(ws, curr_row_num, 1, "", ["vborder"])
        ws.merge_cells(start_row=curr_row_num, start_column=1,
                       end_row=curr_row_num, end_column=16)

    return save_workbook(wb)


def _build_pegawai_lookup(dnp_df: pd.DataFrame) -> dict:
    """Build lookup dictionary for employees by organization code"""
    pegawai_by_org = {}

    # Data untuk direksi
    direksi_data = dnp_df[dnp_df["kode_organisasi"] == "1"]
    if not direksi_data.empty:
        pegawai_by_org["1"] = direksi_data

    # Data untuk organisasi lainnya
    for kode_org in dnp_df["kode_organisasi"].unique():
        if kode_org == "1":
            continue
        org_data = dnp_df[dnp_df["kode_organisasi"].str.startswith(kode_org, na=False)]
        if not org_data.empty:
            pegawai_by_org[kode_org] = org_data

    return pegawai_by_org


def _add_pegawai_row(ws, row_num: int, root_urut: int, child_urut: int, peg) -> None:
    """Add employee data row to worksheet"""
    col_num = itertools.count(start=1)

    # Define column mapping untuk menghindari repetisi
    columns = [
        (root_urut, ["allborder"]),
        (child_urut, ["allborder"]),
        (peg.nama, ["allborder"]),
        (peg.nipam, ["allborder"]),
        (peg.nama_jabatan, ["allborder"]),
        (peg.tmt_jabatan, ["allborder"]),
        (peg.pangkat, ["allborder"]),
        (peg.golongan, ["allborder"]),
        (peg.tmt_golongan, ["allborder"]),
        (peg.mkg_tahun, ["allborder"]),
        (peg.mkg_bulan, ["allborder"]),
        (peg.tmt_kerja, ["allborder"]),
        (peg.mk_tahun, ["allborder"]),
        (peg.mk_bulan, ["allborder"]),
        (peg.pendidikan, ["allborder"]),
        (peg.ttl, ["allborder"])
    ]

    for value, styles in columns:
        cell_builder(ws, row_num, next(col_num), value, styles)
