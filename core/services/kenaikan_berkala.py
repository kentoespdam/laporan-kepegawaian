import io

import pandas as pd
from openpyxl import load_workbook

from core.enums import FilterKenaikanBerkala
from core.excel_helper import font_style, text_align, write_data_to_excel, save_workbook
from core.helper import get_jenis_sk_name, format_date_vectorized, hitung_bulan_vectorized
from core.model.kenaikan_berkala import fetch_kenaikan_berkala


def kenaikan_berkala_data(filter: FilterKenaikanBerkala = FilterKenaikanBerkala.BULAN_INI):
    data = fetch_kenaikan_berkala(filter)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(df: pd.DataFrame):
    date_columns = ["tmt_berlaku", "tmt_kenaikan", "tmt_jabatan", "tmt_golongan", "tmt_kerja", "tanggal_lahir"]
    for col in date_columns:
        if col in df.columns:
            df[col] = format_date_vectorized(df[col])

    df["mkg_bulan"] = hitung_bulan_vectorized(df["mkg_tahun"], df["mkg_bulan"])
    df["mk_bulan"] = hitung_bulan_vectorized(df["mk_tahun"], df["mk_bulan"])

    df["jenis_sk"] = get_jenis_sk_name(df["jenis_sk"])
    return df


def to_excel(title_text: str, filter: FilterKenaikanBerkala) -> io.BytesIO | None:
    """Generate Excel file from kenaikan berkala data"""
    data = kenaikan_berkala_data(filter)
    if data.empty:
        return None

    wb = load_workbook("template/template_kenaikan_berkala.xlsx")
    ws = wb.active

    # Set title
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = title_text
    title_cell.alignment = text_align("center")
    title_cell.font = font_style("bold")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    # Define column mappings for cleaner code
    column_mappings = [
        ("index", lambda idx, row: int(idx + 1), ["bold", "allborder"]),
        ("nama", lambda idx, row: row["nama"], ["allborder"]),
        ("nipam", lambda idx, row: row["nipam"], ["allborder"]),
        ("nama_jabatan", lambda idx, row: row["nama_jabatan"], ["allborder"]),
        ("tmt_jabatan", lambda idx, row: row["tmt_jabatan"], ["allborder"]),
        ("pangkat", lambda idx, row: row["pangkat"], ["allborder"]),
        ("golongan", lambda idx, row: row["golongan"], ["allborder"]),
        ("tmt_golongan", lambda idx, row: row["tmt_golongan"], ["allborder"]),
        ("mkg_tahun", lambda idx, row: row["mkg_tahun"], ["allborder"]),
        ("mkg_bulan", lambda idx, row: row["mkg_bulan"], ["allborder"]),
        ("tmt_kerja", lambda idx, row: row["tmt_kerja"], ["allborder"]),
        ("mk_tahun", lambda idx, row: row["mk_tahun"], ["allborder"]),
        ("mk_bulan", lambda idx, row: row["mk_bulan"], ["allborder"]),
        ("pendidikan_terakhir", lambda idx, row: row["pendidikan_terakhir"], ["allborder"]),
        ("tempat_tanggal_lahir", lambda idx, row: f"{row['tempat_lahir']}, {row['tanggal_lahir']}", ["allborder"]),
        ("empty", lambda idx, row: "", ["allborder"])
    ]

    # Write data to Excel
    write_data_to_excel(ws, data, column_mappings, 7)

    return save_workbook(wb)
