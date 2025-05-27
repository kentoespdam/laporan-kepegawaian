import io
import itertools
import swifter  # noqa: F401
import pandas as pd
from openpyxl import load_workbook

from core.excel_helper import cell_builder, font_style, text_align
from core.model.kontrak import FILTER_KONTRAK, fetch_kontrak


def kontrak_data(filter: FILTER_KONTRAK = FILTER_KONTRAK.AKTIF):
    data = fetch_kontrak(filter)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(pd: pd.DataFrame) -> pd.DataFrame:
    pd["tanggal_mulai"] = pd["tanggal_mulai"].swifter.apply(
        lambda x: x.strftime("%Y-%m-%d"))
    pd["tanggal_selesai"] = pd["tanggal_selesai"].swifter.apply(
        lambda x: x.strftime("%Y-%m-%d"))

    return pd


def to_excel(title_text: str, filter: FILTER_KONTRAK = FILTER_KONTRAK.AKTIF):
    data = kontrak_data(filter)
    if data.empty:
        return None

    wb = load_workbook("template/template_kontrak.xlsx")
    ws = wb.active

    title_cell = ws.cell(row=2, column=1)
    title_cell.value = title_text
    title_cell.alignment = text_align("center")
    title_cell.font = font_style("bold")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    row_num = itertools.count(start=5)
    for index, row in data.iterrows():
        col_num = itertools.count(start=1)
        current_row = next(row_num)
        cell_builder(ws, current_row, next(col_num),
                     int(index+1), ["bold", "allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nipam"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nomor_kontrak"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_organisasi"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_jabatan"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["tanggal_mulai"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["tanggal_selesai"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["sisa_bulan"], ["allborder"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
