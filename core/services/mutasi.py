import io
import itertools

import pandas as pd
import swifter  # noqa: F401
from openpyxl import load_workbook

from core.enums import get_jenis_mutasi_name
from core.excel_helper import cell_builder, font_style, text_align
from core.helper import get_nama_bulan
from core.model.mutasi import fetch_mutasi


def mutasi_data(from_date: str, to_date: str, jenis_mutasi: int = None) -> pd.DataFrame:
    data = fetch_mutasi(from_date, to_date, jenis_mutasi)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(df: pd.DataFrame) -> pd.DataFrame:
    df["jenis_mutasi"] = df["jenis_mutasi"].swifter.apply(
        lambda x: get_jenis_mutasi_name(x))
    df["tmt_berlaku"] = df["tmt_berlaku"].swifter.apply(
        lambda x: x.strftime("%Y-%m-%d") if not pd.isna(x) else None
    )

    return df


def to_excel(from_date: str, to_date: str, jenis_mutasi: int = None) -> io.BytesIO:
    data = mutasi_data(from_date, to_date, jenis_mutasi)
    if data.empty:
        return None

    wb = load_workbook('template/template_mutasi.xlsx')
    ws = wb.active

    title_cell = ws.cell(row=2, column=1)
    title_cell.value = "Periode: {}-{}-{} s/d {}-{}-{}".format(
        from_date[8:10],
        get_nama_bulan(int(from_date[5:7])),
        from_date[0:4],
        to_date[8:10],
        get_nama_bulan(to_date[5:7]),
        to_date[0:4]
    )
    title_cell.font = font_style("bold")
    title_cell.alignment = text_align("center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row_num = itertools.count(start=5)
    for urut, row in data.iterrows():
        col_num = itertools.count(start=1)
        current_row = next(row_num)
        cell_builder(ws, current_row, next(col_num),
                     urut+1, ["bold", "allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["jenis_mutasi"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nipam"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["tmt_berlaku"], ["allborder"])
        cell_builder(ws, current_row, next(col_num), row["nama_organisasi_lama"],
                     ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_jabatan_lama"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_golongan_lama"], ["allborder"])
        cell_builder(ws, current_row, next(col_num), row["nama_organisasi"],
                     ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_jabatan"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["nama_golongan"], ["allborder"])
        cell_builder(ws, current_row, next(col_num),
                     row["notes"], ["allborder"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
