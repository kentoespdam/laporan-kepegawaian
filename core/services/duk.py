import io

import pandas as pd
from openpyxl import load_workbook

from core.excel_helper import write_data_to_excel, save_workbook
from core.helper import hitung_sisa_bulan, format_date_series, get_status_pegawai
from core.model.duk import fetch_duk


def duk_data():
    data = fetch_duk()
    data = cleanup(data)
    return data


def cleanup(df: pd.DataFrame) -> pd.DataFrame:
    df["tmt_golongan"] = format_date_series(df["tmt_golongan"])
    df["tmt_jabatan"] = format_date_series(df["tmt_jabatan"])
    df["tmt_kerja"] = format_date_series(df["tmt_kerja"])
    df["mk_bulan"] = hitung_sisa_bulan(df["mk_tahun"], df["mk_bulan"])
    df["status_pegawai"] = get_status_pegawai(df["status_pegawai"])
    return df


def to_excel(tahun: int, bulan: str) -> io.BytesIO:
    """Generate Excel file from DUK data"""
    data = duk_data()
    wb = load_workbook('template/template_duk.xlsx')
    ws = wb.active

    # Set title
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = f"BULAN : {bulan} {tahun}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)

    # Define column mappings for cleaner code
    column_mappings = [
        ("index", lambda idx, _: int(idx + 1), ["bold", "allborder"]),
        ("nama", lambda _, row: row["nama"], ["allborder"]),
        ("nipam", lambda _, row: row["nipam"], ["allborder"]),
        ("golongan", lambda _, row: row["golongan"], ["allborder"]),
        ("pangkat", lambda _, row: row["pangkat"], ["allborder"]),
        ("tmt_golongan", lambda _, row: row["tmt_golongan"], ["allborder"]),
        ("nama_jabatan", lambda _, row: row["nama_jabatan"], ["allborder"]),
        ("tmt_jabatan", lambda _, row: row["tmt_jabatan"], ["allborder"]),
        ("tmt_kerja", lambda _, row: row["tmt_kerja"], ["allborder"]),
        ("mk_tahun", lambda _, row: row["mk_tahun"], ["allborder"]),
        ("mk_bulan", lambda _, row: row["mk_bulan"], ["allborder"]),
        ("jurusan", lambda _, row: row["jurusan"], ["allborder"]),
        ("tahun_lulus", lambda _, row: row["tahun_lulus"], ["allborder"]),
        ("tingkat_pendidikan", lambda _, row: row["tingkat_pendidikan"], ["allborder"]),
        ("usia", lambda _, row: row["usia"], ["allborder"]),
        ("status_pegawai", lambda _, row: row["status_pegawai"], ["allborder"])
    ]

    # Write data to Excel
    write_data_to_excel(ws, data, column_mappings, 7)

    return save_workbook(wb)
