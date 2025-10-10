from io import BytesIO

import numpy as np
import pandas as pd
from icecream import ic
from openpyxl import load_workbook
from typing import Optional

from app.core.enums import FilterKenaikanBerkala, JenisSk
from app.core.excel_helper import font_style, text_align, write_data_to_excel, save_workbook
from app.core.helper import format_date_vectorized, cleanup_nan_to_zero_safe
from app.models.kenaikan_berkala import KenaikanBerkalaModel


class KenaikanBerkalaService:
    """Service for handling periodic promotion data with Excel export functionality."""

    # Constants for better maintainability
    DATE_COLUMNS = {
        "kenaikan_berikutnya", "tanggal_eksekusi_sanksi", "tanggal_lahir",
        "tmt_berlaku", "tmt_golongan", "tmt_jabatan", "tmt_kerja"
    }

    COLUMN_MAPPINGS = [
        ("index", lambda idx, row: idx + 1, ["bold", "allborder"]),
        ("nama", lambda idx, row: row["nama"], ["allborder"]),
        ("nipam", lambda idx, row: row["nipam"], ["allborder"]),
        ("nama_jabatan", lambda idx, row: row["nama_jabatan"], ["allborder"]),
        ("tmt_jabatan", lambda idx, row: row["tmt_jabatan"], ["allborder"]),
        ("pangkat", lambda idx, row: row["pangkat"], ["allborder"]),
        ("golongan", lambda idx, row: row["golongan"], ["allborder"]),
        ("tmt_golongan", lambda idx, row: row["tmt_golongan"], ["allborder"]),
        ("mkg_tahun", lambda idx, row: row["mkg_tahun"], ["allborder"]),
        ("mkg_bulan", lambda idx, row: row["mkg_bulan"], ["allborder"]),
        ("tmt_kerja", lambda idx, row: row["tmt_kerja"], ["allborder"]),
        ("mk_tahun", lambda idx, row: row["mk_tahun"], ["allborder"]),
        ("mk_bulan", lambda idx, row: row["mk_bulan"], ["allborder"]),
        ("pendidikan_terakhir", lambda idx, row: row["pendidikan_terakhir"], ["allborder"]),
        ("tempat_tanggal_lahir", lambda idx, row: f"{row['tempat_lahir']}, {row['tanggal_lahir']}", ["allborder"]),
        ("empty", lambda idx, row: "", ["allborder"])
    ]

    TEMPLATE_PATH = "template/template_kenaikan_berkala.xlsx"

    def __init__(self):
        self.kbm = KenaikanBerkalaModel()

    def fetch(self,
              filter_type: FilterKenaikanBerkala = FilterKenaikanBerkala.BULAN_INI,
              jenis_sk: JenisSk = JenisSk.SK_KENAIKAN_GAJI_BERKALA) -> pd.DataFrame:
        """Fetch and clean promotion data based on filter and SK type."""
        df = self.kbm.fetch(filter_type, jenis_sk)
        return _cleanup_data(df, jenis_sk) if not df.empty else pd.DataFrame()

    def fetch_count(self,
              filter_type: FilterKenaikanBerkala = FilterKenaikanBerkala.BULAN_INI,
              jenis_sk: JenisSk = JenisSk.SK_KENAIKAN_GAJI_BERKALA) -> pd.DataFrame:
        return self.kbm.fetch_count(filter_type, jenis_sk)

    def to_excel(self,
                 title_text: str,
                 filter_type: FilterKenaikanBerkala = FilterKenaikanBerkala.BULAN_INI,
                 jenis_sk: JenisSk = JenisSk.SK_KENAIKAN_GAJI_BERKALA) -> Optional[bytes]:
        """Generate Excel file from promotion data."""
        df = self.fetch(filter_type, jenis_sk)
        return _generate_excel_report(title_text, df) if not df.empty else None


def _cleanup_data(df: pd.DataFrame, jenis_sk: JenisSk) -> pd.DataFrame:
    """Clean and format the dataframe."""
    result = df.copy()

    # Convert boolean columns
    result["is_pending_gaji"] = result["is_pending_gaji"].eq(b'\x01')
    result["is_pending_pangkat"] = result["is_pending_pangkat"].eq(b'\x01')

    # Clean execution date
    result["tanggal_eksekusi_sanksi"] = _cleanup_execution_date(
        jenis_sk,
        result["tanggal_eksekusi_sanksi"],
        result["is_pending_gaji"],
        result["is_pending_pangkat"]
    )

    result = cleanup_nan_to_zero_safe(result, columns=["mkg_bulan", "mkg_tahun"])

    # Format date columns
    _format_date_columns(result)

    return result


def _cleanup_execution_date(jenis_sk: JenisSk,
                            execution_date: pd.Series,
                            is_pending_gaji: pd.Series,
                            is_pending_pangkat: pd.Series) -> pd.Series:
    """Clean execution date based on SK type and pending status."""
    condition_map = {
        JenisSk.SK_KENAIKAN_GAJI_BERKALA: ~is_pending_gaji,
        JenisSk.SK_KENAIKAN_PANGKAT_GOLONGAN: ~is_pending_pangkat
    }

    condition = condition_map.get(jenis_sk)
    return np.where(condition, None, execution_date) if condition is not None else execution_date


def _format_date_columns(df: pd.DataFrame) -> None:
    """Format date columns in-place."""
    date_columns = [col for col in KenaikanBerkalaService.DATE_COLUMNS if col in df.columns]
    for column in date_columns:
        df[column] = format_date_vectorized(df[column])


def _generate_excel_report(title_text: str, df: pd.DataFrame) -> BytesIO:
    """Generate Excel report from dataframe."""
    wb = load_workbook(KenaikanBerkalaService.TEMPLATE_PATH)
    ws = wb.active

    # Set title
    _set_excel_title(ws, title_text)

    # Write data to Excel
    write_data_to_excel(ws, df, KenaikanBerkalaService.COLUMN_MAPPINGS, 7)

    return save_workbook(wb)


def _set_excel_title(worksheet, title_text: str) -> None:
    """Set and format title in Excel worksheet."""
    title_cell = worksheet.cell(row=2, column=1)
    title_cell.value = title_text
    title_cell.alignment = text_align("center")
    title_cell.font = font_style("bold")
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
