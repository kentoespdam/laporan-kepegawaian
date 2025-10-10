import pandas as pd
from openpyxl import load_workbook

from app.core.enums import FilterLepasTanggunganAnak
from app.core.excel_helper import font_style, text_align, write_data_to_excel, save_workbook
from app.core.helper import format_date_vectorized, cleanup_boolean_dynamic
from app.models.lepas_tanggungan_anak import (
    LepasTanggunganAnakModel,
)


class LepasTanggunganAnakService:
    def __init__(self):
        self.model = LepasTanggunganAnakModel()
        self.filter = FilterLepasTanggunganAnak.BULAN_INI
        self.df = pd.DataFrame()

    def fetch(self, filter: FilterLepasTanggunganAnak = FilterLepasTanggunganAnak.BULAN_INI):
        self.filter = filter
        self._get_data()
        if not self.df.empty:
            self._cleanup()
        return self.df

    def fetch_count(self, filter: FilterLepasTanggunganAnak = FilterLepasTanggunganAnak.BULAN_INI):
        return self.model.count(filter)

    def to_excel(self, title_text: str, filter: FilterLepasTanggunganAnak = FilterLepasTanggunganAnak.BULAN_INI):
        self.filter = filter
        self._get_data()
        if self.df.empty:
            return None
        self._cleanup()

        return self._generate_excel(title_text)

    def _cleanup(self):
        self.df["tanggal_lahir"] = format_date_vectorized(self.df["tanggal_lahir"])
        self.df["tanggungan"] = cleanup_boolean_dynamic(self.df["tanggungan"])

    def _get_data(self):
        self.df = self.model.fetch(self.filter)

    def _generate_excel(self, title_text: str):
        data = self.df
        wb = load_workbook("template/template_lta.xlsx")
        ws = wb.active

        title_cell = ws.cell(row=2, column=1)
        title_cell.value = title_text
        title_cell.alignment = text_align("center")
        title_cell.font = font_style("bold")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

        column_mappings = [
            ("index", lambda idx, _: int(idx + 1), ["bold", "allborder"]),
            ("nama_anak", lambda _, row: row["nama_anak"], ["allborder"]),
            ("jenis_kelamin", lambda _, row: row["jenis_kelamin"], ["allborder"]),
            ("tanggal_lahir", lambda _, row: row["tanggal_lahir"], ["allborder"]),
            ("umur", lambda _, row: row["umur"], ["allborder"]),
            ("status_pendidikan", lambda _, row: row["status_pendidikan"], ["allborder"]),
            ("nama_karyawan", lambda _, row: row["nama_karyawan"], ["allborder"]),
            ("nipam", lambda _, row: row["nipam"], ["allborder"]),
            ("nama_jabatan", lambda _, row: row["nama_jabatan"], ["allborder"]),
        ]

        # Write data to Excel
        write_data_to_excel(ws, data, column_mappings, 5)

        return save_workbook(wb)
