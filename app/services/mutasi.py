import io
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from app.core.excel_helper import font_style, text_align, save_workbook, write_data_to_excel
from app.core.helper import get_nama_bulan, get_jenis_mutasi_name, format_date_vectorized
from app.models.mutasi import fetch_mutasi


def mutasi_data(from_date: str, to_date: str, jenis_mutasi: int = None) -> pd.DataFrame:
    data = fetch_mutasi(from_date, to_date, jenis_mutasi)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(df: pd.DataFrame) -> pd.DataFrame:
    df["jenis_mutasi"] = get_jenis_mutasi_name(df["jenis_mutasi"])
    df["tmt_berlaku"] = format_date_vectorized(df["tmt_berlaku"])

    return df


def _format_period_title_fast(from_date: str, to_date: str) -> str:
    """Fast version of period title formatting"""
    # Use slicing with variables to avoid repeated slicing
    from_day, from_month_num, from_year = from_date[8:10], from_date[5:7], from_date[0:4]
    to_day, to_month_num, to_year = to_date[8:10], to_date[5:7], to_date[0:4]

    # Cache bulan names if called frequently
    bulan_cache = {}
    from_bulan = bulan_cache.get(from_month_num) or get_nama_bulan(int(from_month_num))
    to_bulan = bulan_cache.get(to_month_num) or get_nama_bulan(int(to_month_num))

    return f"Periode: {from_day}-{from_bulan}-{from_year} s/d {to_day}-{to_bulan}-{to_year}"


def to_excel(from_date: str, to_date: str, jenis_mutasi: int = None) -> Optional[io.BytesIO]:
    """Optimized version for better performance with large datasets"""
    data = mutasi_data(from_date, to_date, jenis_mutasi)
    if data.empty:
        return None

    wb = load_workbook('template/template_mutasi.xlsx')
    ws = wb.active

    # Set title (optimized)
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = _format_period_title_fast(from_date, to_date)
    title_cell.font = font_style("bold")
    title_cell.alignment = text_align("center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # Pre-define columns for faster access
    columns = [
        ("jenis_mutasi", lambda idx, _: int(idx + 1), ["bold", "allborder"]),
        ("nipam", lambda _, row: row["nipam"], ["allborder"]),
        ("nama", lambda _, row: row["nama"], ["allborder"]),
        ("tmt_berlaku", lambda _, row: row["tmt_berlaku"], ["allborder"]),
        ("nama_organisasi_lama", lambda _, row: row["nama_organisasi_lama"], ["allborder"]),
        ("nama_jabatan_lama", lambda _, row: row["nama_jabatan_lama"], ["allborder"]),
        ("nama_golongan_lama", lambda _, row: row["nama_golongan_lama"], ["allborder"]),
        ("nama_organisasi", lambda _, row: row["nama_organisasi"], ["allborder"]),
        ("nama_jabatan", lambda _, row: row["nama_jabatan"], ["allborder"]),
        ("nama_golongan", lambda _, row: row["nama_golongan"], ["allborder"]),
        ("notes", lambda _, row: row["notes"], ["allborder"]),
    ]

    write_data_to_excel(ws, data, columns, 5)

    return save_workbook(wb)
