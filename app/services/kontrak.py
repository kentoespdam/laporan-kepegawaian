import pandas as pd
from openpyxl import load_workbook

from app.core.excel_helper import font_style, text_align, write_data_to_excel, save_workbook
from app.core.helper import format_date_vectorized
from app.models.kontrak import FilterKontrak, fetch_kontrak


def kontrak_data(filter: FilterKontrak = FilterKontrak.AKTIF):
    data = fetch_kontrak(filter)
    if not data.empty:
        data = _cleanup(data)
    return data


def _cleanup(pd: pd.DataFrame) -> pd.DataFrame:
    pd["tanggal_mulai"] = format_date_vectorized(pd["tanggal_mulai"])
    pd["tanggal_selesai"] = format_date_vectorized(pd["tanggal_selesai"])
    return pd


def to_excel(title_text: str, filter: FilterKontrak = FilterKontrak.AKTIF):
    data = kontrak_data(filter)
    if data.empty:
        return None

    wb = load_workbook("template/template_kontrak.xlsx")
    ws = wb.active

    title_cell = ws.cell(row=2, column=1)
    title_cell.value = title_text
    title_cell.alignment = text_align("center")
    title_cell.font = font_style("bold")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    column_mappings = [
        ("index", lambda idx, _: int(idx + 1), ["bold", "allborder"]),
        ("nipam", lambda _, row: row["nipam"], ["allborder"]),
        ("nama", lambda _, row: row["nama"], ["allborder"]),
        ("nomor_kontrak", lambda _, row: row["nomor_kontrak"], ["allborder"]),
        ("nama_organisasi", lambda _, row: row["nama_organisasi"], ["allborder"]),
        ("nama_jabatan", lambda _, row: row["nama_jabatan"], ["allborder"]),
        ("tanggal_mulai", lambda _, row: row["tanggal_mulai"], ["allborder"]),
        ("tanggal_selesai", lambda _, row: row["tanggal_selesai"], ["allborder"]),
        ("sisa_bulan", lambda _, row: row["sisa_bulan"], ["allborder"]),
    ]

    # Write data to Excel
    write_data_to_excel(ws, data, column_mappings, 5)

    return save_workbook(wb)
