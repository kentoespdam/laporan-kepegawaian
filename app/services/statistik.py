import io
from typing import Optional

import pandas as pd
from openpyxl.reader.excel import load_workbook

from app.core.enums import get_status_pegawai_name
from app.core.excel_helper import cell_builder, write_data_to_excel, save_workbook
from app.core.helper import get_agama, get_nama_bulan
from app.models.statistik import (
    fetch_by_agama,
    fetch_by_gelar,
    fetch_by_golongan,
    fetch_by_jenis_kelamin,
    fetch_by_pendidikan_1,
    fetch_by_status_pegawai,
    fetch_by_umur, fetch_by_pendidikan_2,
)


def fetch_golongan_data():
    data = fetch_by_golongan()
    data = data.astype({"jml_l": int, "jml_p": int, "total": int})
    data["persen"] = round(((data["total"] / data["total"].sum()) * 100), 2)
    return data


def fetch_pendidikan_1_data():
    data = fetch_by_pendidikan_1()
    data = data.astype({"total": int})
    data["persen"] = round(((data["total"] / data["total"].sum()) * 100), 2)
    return data


def fetch_pendidikan_2_data(tahun: int, bulan: int):
    data = fetch_by_pendidikan_2(tahun, bulan)
    return data


def generate_excel_pendidikan_2(tahun: int, bulan: int) -> Optional[io.BytesIO]:
    data = fetch_by_pendidikan_2(tahun, bulan)
    if data.empty:
        return None

    total_data = data["id"].size

    wb = load_workbook("template/template_pendidikan_2.xlsx")
    ws = wb.active

    # Set title
    title_cell = ws.cell(row=2, column=1)
    title_cell.value = f"BULAN : {get_nama_bulan(bulan)} {tahun}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)

    # Define column mappings
    column_mappings = [
        ("pendidikan", lambda _, row: row["pendidikan"], ["allborder"]),
        ("non_golongan", lambda _, row: row["non_golongan"], ["allborder"]),
        ("golongan_a", lambda _, row: row["golongan_a"], ["allborder"]),
        ("golongan_b", lambda _, row: row["golongan_b"], ["allborder"]),
        ("golongan_c", lambda _, row: row["golongan_c"], ["allborder"]),
        ("golongan_d", lambda _, row: row["golongan_d"], ["allborder"]),
        ("jml_golongan", lambda _, row: row["jml_golongan"], ["allborder"]),
        ("kontrak", lambda _, row: row["kontrak"], ["allborder"]),
        ("capeg", lambda _, row: row["capeg"], ["allborder"]),
        ("honorer", lambda _, row: row["honorer"], ["allborder"]),
        ("tetap", lambda _, row: row["tetap"], ["allborder"]),
        ("jml_status_pegawai", lambda _, row: row["jml_status_pegawai"], ["allborder"]),
        ("adm", lambda _, row: row["adm"], ["allborder"]),
        ("pelayanan", lambda _, row: row["pelayanan"], ["allborder"]),
        ("teknik", lambda _, row: row["teknik"], ["allborder"]),
        ("jml_unit_kerja", lambda _, row: row["jml_unit_kerja"], ["allborder"]),
        ("pria", lambda _, row: row["pria"], ["allborder"]),
        ("wanita", lambda _, row: row["wanita"], ["allborder"]),
        ("jml_jenis_kelamin", lambda _, row: row["jml_jenis_kelamin"], ["allborder"]),
    ]
    start_row = 6
    write_data_to_excel(ws, data, column_mappings, start_row)
    start_footer_row = start_row + total_data
    max_data_row_num = start_footer_row - 1

    formulas = [
        "JML",
        f"=SUM(B{start_row}:B{max_data_row_num})",
        f"=SUM(C{start_row}:C{max_data_row_num})",
        f"=SUM(D{start_row}:D{max_data_row_num})",
        f"=SUM(E{start_row}:E{max_data_row_num})",
        f"=SUM(F{start_row}:F{max_data_row_num})",
        f"=SUM(G{start_row}:G{max_data_row_num})",
        f"=SUM(H{start_row}:H{max_data_row_num})",
        f"=SUM(I{start_row}:I{max_data_row_num})",
        f"=SUM(J{start_row}:J{max_data_row_num})",
        f"=SUM(K{start_row}:K{max_data_row_num})",
        f"=SUM(L{start_row}:L{max_data_row_num})",
        f"=SUM(M{start_row}:M{max_data_row_num})",
        f"=SUM(N{start_row}:N{max_data_row_num})",
        f"=SUM(O{start_row}:O{max_data_row_num})",
        f"=SUM(P{start_row}:P{max_data_row_num})",
        f"=SUM(Q{start_row}:Q{max_data_row_num})",
        f"=SUM(R{start_row}:R{max_data_row_num})",
        f"=SUM(S{start_row}:S{max_data_row_num})",
    ]
    for index, formula in enumerate(formulas, start=1):
        cell_builder(ws, start_footer_row, index, formula, ["allborder", "bold"])

    return save_workbook(wb)


def fetch_umur_data():
    data = fetch_by_umur()
    data = data.astype({"total": int})
    data["persen"] = round(((data["total"] / data["total"].sum()) * 100), 2)
    data_range = _generate_data_umur2(data)
    data_range["persen"] = round(((data_range["total"] / data_range["total"].sum()) * 100), 2)
    return data, data_range


def _generate_data_umur2(df: pd.DataFrame) -> pd.DataFrame:
    list_range = ["<20", "20-29", "30-39", "40-49", "50-59", ">60"]
    range1 = df.query("umur < 20")["total"].sum()
    range2 = df.query("umur >= 20 and umur < 30")["total"].sum()
    range3 = df.query("umur >= 30 and umur < 40")["total"].sum()
    range4 = df.query("umur >= 40 and umur < 50")["total"].sum()
    range5 = df.query("umur >= 50 and umur < 60")["total"].sum()
    range6 = df.query("umur >= 60")["total"].sum()
    total = [range1, range2, range3, range4, range5, range6]
    return pd.DataFrame({"range": list_range, "total": total})


def fetch_jenis_kelamin_data():
    data = fetch_by_jenis_kelamin()
    data = data.astype({"total": int})
    data["persen"] = round(((data["total"] / data["total"].sum()) * 100), 2)
    return data


def fetch_gelar_data():
    data = fetch_by_gelar()
    data = data.astype({"total": int})
    data["persen"] = (data["total"] / data["total"].sum()) * 100
    return data


def fetch_agama_data():
    data = fetch_by_agama()
    data["agama"] = data["agama"].apply(lambda x: get_agama(x))
    data = data.astype({"total": int})
    data["persen"] = round(((data["total"] / data["total"].sum()) * 100), 2)
    return data


def fetch_status_pegawai_data():
    data = fetch_by_status_pegawai()
    data["status_pegawai"] = data["status_pegawai"].apply(
        lambda x: get_status_pegawai_name(x)
    )
    data = data.astype({"total": int})
    data["persen"] = (data["total"] / data["total"].sum()) * 100
    # sort data by status_pegawai
    data = data.sort_values(by="status_pegawai", ascending=False)
    return data
